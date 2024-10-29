import os
from openpyxl import load_workbook
import pandas as pd
from PySide6.QtWidgets import QFileDialog, QDialog
from view.mail_dialog import MailDialog

def define_quantity(row):
    qte = int(row['Qté'])
    colisage = int(row['Colisage'])
    if colisage != 1:
        qte = qte / colisage
    return qte

def gw(controller):
    pathDesktop = controller.localSettings.get("path_desktop")
    priceFilePath = controller.file_present(pathDesktop, "Liste de prix")
    if not priceFilePath:
        priceFilePath, _ = QFileDialog.getOpenFileName(filter=("XLS (*.xlsx)"), dir=pathDesktop, caption="Sélectionner la liste de prix")
    try : 
        priceList = pd.read_excel(
            priceFilePath, 
            sheet_name=0, 
            header=None, 
            usecols=[5, 9],
            dtype={5: str}
        ).dropna(axis=0, how="all")
        priceList = priceList.iloc[1:]
        priceList.rename(columns={5:"Référence", 9: "Colisage"}, inplace=True)
        priceList["Référence"] = priceList["Référence"].astype(str)
    except FileNotFoundError:
        return

    exportFilePath, _ = QFileDialog.getOpenFileName(filter=("CSV (*.csv)"), dir=pathDesktop, caption="Sélectionner l'export")
    try : 
        export = pd.read_csv(
            exportFilePath, 
            sep=";", 
            usecols=["Référence", "Qté"], 
            dtype={"Référence": str}
        ).dropna(how="all")
        export["Référence"] = export["Référence"].astype(str)
    except FileNotFoundError:
        return

    results = pd.merge(priceList, export, on='Référence', how='inner')
    unfoundRefs = export[~export['Référence'].isin(results['Référence'])]
    unfoundRefs = unfoundRefs["Référence"].tolist()
    results['Qté'] = results.apply(define_quantity, axis=1)
    print(", ".join(unfoundRefs))

    wb = load_workbook(priceFilePath)
    ws = wb[wb.sheetnames[0]]
    quantities = dict(zip(results["Référence"], results["Qté"]))
    for row in ws.iter_rows(max_col=ws.max_column):
        ref = row[5].value
        if ref in quantities:
            row[6].value = quantities[ref]

    filePath = os.path.join(pathDesktop, "Commande.xlsx")
    wb.save(filePath)
    wb.close()

    fromAdress=controller.globalSettings.get("gw_from", "")
    toAdress = controller.globalSettings.get("gw_to", "")
    name = toAdress.split(".")[0]
    subject = "Les 7 Royaumes - Nouvelle Commande CDF-XXXX"
    body = (
        f"Bonjour {name.title()} :)\n\n"
        f"Voici en pièce jointe, ma commande de réassort & nouveautés.\n"
        f"Merci d'en accuser bonne réception.\n"
        f"Si besoin je suis disponible au 0476545835 ou en réponse à ce mail pour en discuter.\n\n"
        f"Bien cordialement, Hugo.\n- Force & Honneur !"
    )
    attachments = [filePath]

    mailDialog = MailDialog(subject, body, attachments, fromAdress, toAdress, controller.mainWindow)
    if mailDialog.exec() == QDialog.Accepted:
        # Récupérer les valeurs mises à jour
        subject, body, fromAdress, toAdress = mailDialog.get_mail_content()
        controller.mail.send_email(
            fromAdress=fromAdress,
            toAdress=toAdress,
            subject=subject,
            body=body,
            attachments=attachments
        )

    controller.mail.send_email(
        fromAdress=controller.globalSettings.get("gw_from", ""),
        toAdress=controller.globalSettings.get("gw_errors_to", ""),
        subject="Erreur références GW",
        body="Bonjour,\nLes références suivantes n'ont pas été trouvées: \n -" + "\n -".join(unfoundRefs)
    )
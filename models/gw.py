import pandas as pd
from openpyxl import load_workbook

def define_quantity(row):
    qte = int(row['Qté'])
    colisage = int(row['Colisage'])
    if colisage != 1:
        qte = qte / colisage
    return qte

priceList = pd.read_excel(priceFilePath, sheet_name=0, header=None, usecols=[5, 9]).dropna(axis=0, how="all")
priceList = priceList.iloc[1:]
priceList.rename(columns={5:"Référence", 9: "Colisage"}, inplace=True)

export = pd.read_csv(exportFilePath, sep=";", usecols=["Référence", "Qté"]).dropna(how="all")

results = pd.merge(priceList, export, on='Référence', how='inner')
unfoundRefs = export[~export['Référence'].isin(results['Référence'])]
unfoundRefs = unfoundRefs["Référence"].tolist()
results['Qté'] = results.apply(define_quantity, axis=1)

wb = load_workbook(pricefilePath)
ws = wb[wb.sheetnames[0]]
quantities = dict(zip(results["Référence"], results["Qté"]))
for row in ws.iter_rows(max_col=ws.max_column):
    ref = row[5].value
    if ref in quantities:
        row[6].value = quantities[ref]

wb.save("./essai.xlsx")

smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_user = 'gaelle.rem@gmail.com'
smtp_password = 'mdp'

msg = MIMEMultipart()
msg['From'] = smtp_user
msg['To'] = 'coryismyangel@gmail.com'
msg['Subject'] = 'essai'

body = 'Ceci est le corps de l\'email'
msg.attach(MIMEText(body, 'plain'))

try:
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        print("email envoyé avec succès")
except Exception as e:
    print(f'erreur lors de l\'envoi de l\'email : {e}')
